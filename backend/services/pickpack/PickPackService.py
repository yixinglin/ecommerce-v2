import hashlib
from io import BytesIO
from typing import List
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from sp_api.base import Marketplaces
from starlette.responses import StreamingResponse
import utils.time as time_utils
import utils.utilpdf as utilpdf
import utils.stringutils as stringutils
from core.config import settings
from core.db import RedisDataManager, OrderQueryParams
from core.log import logger
from models.convert import convert_to_standard_shipment
from models.orders import StandardOrder
from models.pickpack import BatchOrderConfirmEvent
from models.shipment import StandardShipment
from schemas import ResponseSuccess
from schemas.carriers import PickSlipItemVO
from services.amazon.AmazonService import AmazonService, AmazonOrderService
from services.amazon.bulkOrderService import AmazonBulkPackSlipDE
from services.gls.GlsShipmentService import GlsShipmentService


class PickPackService:
    def __init__(self):
        super().__init__()
        self.gls_service = GlsShipmentService(key_index=settings.GLS_ACCESS_KEY_INDEX)
        self.redis_manager = RedisDataManager()

    def __enter__(self):
        self.gls_service.__enter__()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.gls_service.__exit__(exc_type, exc_val, exc_tb)

    def __generate_order_key(self, orderItems: List[PickSlipItemVO], encode=False):
        items = sorted(orderItems, key=lambda x: x.sku)
        contents = []
        for item in items:
            contents.append(f"{item.quantity}x[{item.sku}]")
        order_key = ";".join(contents)
        if encode:
            md5_hash = hashlib.md5(order_key.encode('utf-8')).hexdigest()
            order_key = f"{item.quantity:03d}_{item.sku[:8]}_{md5_hash[20:36]}"
        for item in items:
            item.orderKey = order_key
        return order_key

    def get_pick_items(self, orders: List[StandardOrder]) -> List[PickSlipItemVO]:
        """
        Get all pick items from the orders.
        :param  orders: List of orders to be picked
        :return:
        """
        pickItems = []
        for i, order in enumerate(orders):
            items = []
            for orderline in order.items:
                pickItem = PickSlipItemVO(
                    orderId=order.orderId,
                    taskId=i + 1,
                    sku=orderline.sku,
                    quantity=orderline.quantity,
                    title=orderline.name,
                    storageLocation="",
                    purchasedAt=order.purchasedAt,
                    zipCode=order.shipAddress.zipCode,
                    city=order.shipAddress.city,
                )
                items.append(pickItem)
            self.__generate_order_key(items, encode=True)
            pickItems.extend(items)
        # Sort by order id and sku
        # pickItems = sorted(pickItems, key=lambda x: (x.orderId, x.sku))
        return pickItems

    def make_pick_slip(self, orders: List[StandardOrder]) -> List[PickSlipItemVO]:
        """
        Generate a pick slip. The pick slip is aggregated by sku and title
        :param orders: List of orders to be picked
        :return: List of PickSlipItemVO
        """
        pickItems = self.get_pick_items(orders)
        items = map(lambda x: x.dict(), pickItems)
        df = pd.DataFrame.from_dict(items)
        df_slip = df.groupby(['sku', 'title'], as_index=False) \
            .agg({'quantity': 'sum', 'sku': 'first', 'orderKey': 'first'})
        df_slip.sort_values(by=['sku', 'title'], inplace=True)
        # Convert to list of dict
        slips = df_slip.to_dict(orient='records')
        # Convert to list of PickSlipItemVO
        slips = list(map(lambda x: PickSlipItemVO(**x), slips))
        return slips

    def sort_packing_orders(self, orders: List[StandardOrder]) -> List[str]:
        """
        Sort the packing orders by the number of orderlines, orderKey and orderId
        :param  orders: List of orders to be packed
        :return: List of orderIds
        """
        pickItems = self.get_pick_items(orders)
        # Sort by [num_orderlines, orderKey and orderId]
        line_counts = stringutils.count_elements(map(lambda x: x.orderId, pickItems))
        sku_counts = stringutils.count_elements(map(lambda x: x.sku, pickItems))
        items = list(sorted(pickItems, key=lambda x: (line_counts[x.orderId], x.orderKey, x.orderId)))
        sorted_refs = list(map(lambda x: x.orderId, items))
        return stringutils.remove_duplicates(sorted_refs)

    def __get_excel_row_color_by_task_id(self, task_id: int = 0):
        gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        white_fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
        if task_id % 2 == 0:
            return white_fill
        else:
            return gray_fill

    def __add_excel_style(self, ws, task_id_col_index=None):
        # 定义边框样式
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        alignment = Alignment(vertical="top", wrap_text=True)
        for row in range(2, ws.max_row + 1):  # 从第二行开始，因为第一行是表头
            if task_id_col_index is None:
                fill = self.__get_excel_row_color_by_task_id(row)
            else:
                task_id = int(ws.cell(row=row, column=task_id_col_index).value)
                fill = self.__get_excel_row_color_by_task_id(task_id)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = alignment

        row_height = 50
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = row_height

    def pick_slip_to_excel(self, orders: List[StandardOrder]) -> bytes:
        """
        Generate a manifest file in excel format for picking orders.
        :param orders: List of orders to be picked
        :return: Bytes of the manifest file in excel format.
        """
        pickItems = self.make_pick_slip(orders)
        items = [o.dict() for o in pickItems]
        df_pick_slip = pd.DataFrame.from_dict(items)
        # Reduce columes
        df_pick_slip = df_pick_slip[['sku', 'quantity', 'storageLocation', 'title']]
        # Sort by sku
        df_pick_slip.sort_values(by=['sku'], inplace=True)
        # Convert dataframe to excel bytes
        with BytesIO() as excel_bytes:
            df_pick_slip.to_excel(excel_bytes, index=False, sheet_name="pick_slip")
            excel_bytes.seek(0)  # 重置文件指针
            wb = load_workbook(excel_bytes)
            ws = wb.active

            column_widths = [33, 10, 20, 80]  # 根据需要调整每列的宽度
            for col_num, width in enumerate(column_widths, start=1):
                col_letter = ws.cell(row=1, column=col_num).column_letter
                ws.column_dimensions[col_letter].width = width

            self.__add_excel_style(ws)

            # 将修改后的工作簿写回BytesIO对象
            new_excel_bytes = BytesIO()
            wb.save(new_excel_bytes)
            new_excel_bytes.seek(0)
            return new_excel_bytes.read()

    def pack_slips_to_excel(self, orders: List[StandardOrder]) -> bytes:
        """
        Generate a manifest file in excel format for packing orders.
        :param orders: List of orders to be packed
        :return: Bytes of the manifest file in excel format.
        """
        items = self.get_pick_items(orders)

        items = [o.dict() for o in items]
        df_order_lines = pd.DataFrame.from_dict(items)
        # Combine city and zip code to a single column
        df_order_lines['city'] = df_order_lines['zipCode'] + " " + df_order_lines['city']

        # Reduce columes
        df_order_lines = df_order_lines[["purchasedAt", 'taskId', 'orderId', 'sku', 'quantity', 'title', "city"]]
        # Format taskIds
        df_order_lines['taskId'] = df_order_lines['taskId'].apply(lambda x: f"{x: 03d}")
        # Format purchasedAt,
        df_order_lines['purchasedAt'] = df_order_lines['purchasedAt'] \
            .apply(lambda x: f"{time_utils.datetime_to_date(x, target_pattern=r'%d.%m.%Y')}")

        with BytesIO() as excel_bytes:
            df_order_lines.to_excel(excel_bytes, index=False, sheet_name="pack_orders")
            excel_bytes.seek(0)  # 重置文件指针
            wb = load_workbook(excel_bytes)
            ws = wb.active

            column_widths = [13, 8, 25, 30, 10, 80, 30]  # 根据需要调整每列的宽度
            for col_num, width in enumerate(column_widths, start=1):
                col_letter = ws.cell(row=1, column=col_num).column_letter
                ws.column_dimensions[col_letter].width = width

            self.__add_excel_style(ws, task_id_col_index=2)

            # 将修改后的工作簿写回BytesIO对象
            new_excel_bytes = BytesIO()
            wb.save(new_excel_bytes)
            new_excel_bytes.seek(0)
            return new_excel_bytes.read()

    def bulk_shipment_for_orders(self, orders: List[StandardOrder],
                                 batchId: str, carrier: str,
                                 sort_by_order_key: bool = False) -> BatchOrderConfirmEvent:
        """
        Bulk create shipments for orders, and store the batch-data in Redis cache.
        :param orders:  List of orders to be shipped
        :param batchId:   Batch Id
        :param carrier:   Carrier name
        :param sort_by_order_key:   Whether to sort the orders by orderKey
        :return:
        """
        logger.info(f"Batch {batchId} to create of {len(orders)} orders.")
        map_orders = {o.orderId: o for o in orders}
        shipments: List[StandardShipment] \
            = list(map(lambda x: convert_to_standard_shipment(x, carrier), orders))
        logger.info(f"Batch {batchId} to create of {len(shipments)} shipments.")
        svc_carrier = None
        if carrier == "gls":
            svc_carrier = self.gls_service
            new_ids, exist_ids = svc_carrier.save_shipments(shipments)

        else:
            raise RuntimeError(f"Carrier {carrier} is not supported yet.")

        logger.info(f"New shipment ids: {new_ids}")
        logger.info(f"Exist shipment ids: {exist_ids}")
        logger.info(f"#Shipments already exist: {len(exist_ids) + len(new_ids)}")
        message = f"Batch {batchId} created.\n" \
                  f"{len(new_ids)} labels created, and {len(exist_ids)} labels already exist. " \
                  f"{len(exist_ids) + len(new_ids)} shipment(s) stored in total."
        orderIds = new_ids + exist_ids
        orders = [map_orders[o] for o in orderIds]

        if sort_by_order_key:
            orderIds = self.sort_packing_orders(orders)
            logger.info(f"# Sorted orderIds: {len(orderIds)}")

        # Sort shipments by orderIds
        shipments = svc_carrier.find_shipments_by_ids(orderIds)
        # TODO: 这里可能有问题，会导致重复或者缺漏某些订单。
        label_pdf = svc_carrier.get_bulk_shipments_labels(orderIds)
        shipmentIds = []
        for ship in shipments:
            tid = ";".join([p.parcelNumber for p in ship.parcels])
            shipmentIds.append(tid)

        # Save batch-data to Redis Cache
        createdAt = time_utils.now()
        event = BatchOrderConfirmEvent(
            createdAt=createdAt,
            taskId=batchId,
            batchId=batchId,
            orderIds=orderIds,
            shipmentIds=shipmentIds,
            message=message,
            parcelLabelB64=utilpdf.pdf_to_str(label_pdf),
        )

        self.cache_batch_event(event)
        return event

    def get_batch_order_confirm_event(self, batchId: str) -> BatchOrderConfirmEvent:
        """
        Get the batch-data from Redis cache.
        :param batchId:   Batch Id
        :return: BatchOrderConfirmEvent
        """
        man = RedisDataManager()
        event_dict = man.get_json(batchId)
        if event_dict is None:
            return None
        return BatchOrderConfirmEvent(**event_dict)

    def generate_batch_id(self, prefix: str = "Any") -> str:
        current_time = time_utils.now(pattern='%Y%m%dT%H%M%SZ')
        return f"BAT:{prefix}-{current_time}"

    def cache_batch_event(self, event: BatchOrderConfirmEvent):
        man = RedisDataManager()
        TIME_TO_LIVE_SEC = 60 * 60 * 24 * 14  # 14 day
        man.set_json(event.batchId, event.dict(), time_to_live_sec=TIME_TO_LIVE_SEC)



class AmazonPickPackService(PickPackService):

    def __init__(self, key_index: int = 0, marketplace: Marketplaces = Marketplaces.DE):
        super().__init__()
        self.amazon_service = AmazonService(key_index=key_index, marketplace=marketplace)
        self.order_service = self.amazon_service.order_service

    def __enter__(self):
        super().__enter__()
        self.amazon_service.__enter__()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        super().__exit__(exc_type, exc_val, exc_tb)
        self.amazon_service.__exit__(exc_type, exc_val, exc_tb)

    def download_batch_pick_slip_excel(self, orderIds: List[str]) -> StreamingResponse:
        orders = self.order_service.find_orders_by_ids(orderIds)
        excel_bytes = self.pick_slip_to_excel(orders)
        filename = f"batch_pick_slip_{time_utils.now(pattern='%Y%m%d_%H%M')}"
        filesize = len(excel_bytes)
        headers = {'Content-Disposition': f'inline; filename="{filename}.xlsx"', 'Content-Length': str(filesize)}
        return StreamingResponse(BytesIO(excel_bytes),
                                 media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                 headers=headers)

    def download_pack_slip_excel(self, orderIds: List[str]) -> StreamingResponse:
        orders = self.order_service.find_orders_by_ids(orderIds)
        refs = self.sort_packing_orders(orders)
        orders = self.order_service.find_orders_by_ids(refs)
        excel_bytes = self.pack_slips_to_excel(orders)
        filename = f"pack_slip_{time_utils.now(pattern='%Y%m%d_%H%M')}"
        filesize = len(excel_bytes)
        headers = {'Content-Disposition': f'inline; filename="{filename}.xlsx"', 'Content-Length': str(filesize)}
        return StreamingResponse(BytesIO(excel_bytes),
                                 media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                 headers=headers)

    def download_all_orders_excel(self, query: OrderQueryParams) -> StreamingResponse:
        orders = self.order_service.find_orders_by_query_params(query)
        if len(orders) == 0:
            logger.info("No orders found.")
            return ResponseSuccess(data=[], size=0, message="No orders found.")
        refs = stringutils.remove_duplicates([order.orderId for order in orders])
        return self.download_pack_slip_excel(refs)

    def bulk_gls_shipments_by_references(self, refs: List[str], carrier: str):
        refs = stringutils.remove_duplicates(refs)
        orders = self.order_service.find_orders_by_ids(refs)
        # Filter out orders that need transparency code
        orders = [o for o in orders if not self.order_service.mdb.need_transparency_code(o)]
        # Number of orders to be shipped
        num_orders = len(orders)
        batchId = self.generate_batch_id("AMZ")
        batchEvent = self.bulk_shipment_for_orders(orders=orders, batchId=batchId, carrier=carrier,
                                                   sort_by_order_key=True)
        # Create pack slip for the batch using the original template of Amazon
        packSlip = AmazonBulkPackSlipDE.add_packslip_to_container(orderIds=batchEvent.orderIds)
        batchEvent.packSlipB64 = stringutils.base64_encode_str(packSlip)
        self.cache_batch_event(batchEvent)
        return {
            "batchId": batchEvent.batchId,
            "orderIds": batchEvent.orderIds,
            "trackIds": batchEvent.shipmentIds,
            "message": f"GLS shipments of {num_orders} orders created successfully.\n" + batchEvent.message,
            "length": len(batchEvent.orderIds)
        }
