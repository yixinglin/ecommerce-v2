from io import BytesIO
from typing import List, Tuple
import pandas as pd
import hashlib

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment

from core.db import MongoDBDataManager, OrderMongoDBDataManager, ShipmentMongoDBDataManager
from models.shipment import StandardShipment
from rest.common.DataManager import CommonMongoDBManager
from schemas.basic import ExternalService
from utils.stringutils import count_elements, remove_duplicates
from utils.time import datetime_to_date
from vo.carriers import PickSlipItemVO


class PickPackMongoDBManager(CommonMongoDBManager):
    def __init__(self, db_host, db_port):
        super().__init__(db_host, db_port)

    def __get_managers_by_reference_format(self, ref) \
            -> Tuple[OrderMongoDBDataManager, MongoDBDataManager, ShipmentMongoDBDataManager]:
        external_service = self.recognize_by_reference_format(ref)
        catalogManager = None
        orderManager = None
        carrierManager = None
        if external_service == ExternalService.Amazon:
            catalogManager = self.amazonCatalogManager
            orderManager = self.amazonDataManager
            carrierManager = self.glsShipmentDataManager
        return orderManager, catalogManager, carrierManager

    def __generate_order_key(self, orderItems: List[PickSlipItemVO], encode=False):
        items = sorted(orderItems, key=lambda x: x.sku)
        contents = []
        for item in items:
            contents.append(f"{item.quantity}x[{item.sku}]")
        order_key = ";".join(contents)
        if encode:
            md5_hash = hashlib.md5(order_key.encode('utf-8')).hexdigest()
            order_key = f"{item.quantity:03d}_{item.sku[:8]}_{md5_hash[20:36]}"
        for item in items:
            item.orderKey = order_key
        return order_key

    def find_unshipped_amazon_orders(self):
        """
        TODO: Find all unshipped orders
        :return:
        """
        orders = self.amazonDataManager.find_unshipped_orders(days_ago=7, up_to_date=False)
        return orders

    def find_shipments_by_id(self, ids: List[str]) -> List[StandardShipment]:
        orderManager, catalogManager, carrierManager \
            = self.__get_managers_by_reference_format(ids[0])
        shipments = list(carrierManager.find_shipments_by_ids(ids))
        return shipments


    def get_pick_items_by_references(self, refs: List[str]) -> List[PickSlipItemVO]:
        """
        TODO: Get all pick items by references of orders.
        :param refs: List of references
        :return:
        """
        refs =remove_duplicates(refs)
        orderManager, _, _ = self.__get_managers_by_reference_format(refs[0])
        orders = orderManager.find_orders_by_ids(refs)

        pickItems = []
        for i, order in enumerate(orders):
            items = []
            for orderline in order.items:
                pickItem = PickSlipItemVO(
                    orderId=order.orderId,
                    taskId=i + 1,
                    sku=orderline.sku,
                    quantity=orderline.quantity,
                    title=orderline.name,
                    storageLocation="",
                    purchasedAt=order.purchasedAt,
                    zipCode=order.shipAddress.zipCode,
                    city=order.shipAddress.city,
                )
                items.append(pickItem)
            self.__generate_order_key(items, encode=True)
            pickItems.extend(items)
        # Sort by order id and sku
        # pickItems = sorted(pickItems, key=lambda x: (x.orderId, x.sku))
        return pickItems

    def make_pick_slip_with_references(self, refs: List[str]) -> List[PickSlipItemVO]:
        """
        TODO: Get a pick slip by references. The pick slip is aggregated by sku and title
        :param refs: List of references
        :return:
        """
        pickItems = self.get_pick_items_by_references(refs)
        items = map(lambda x: x.dict(), pickItems)
        df = pd.DataFrame.from_dict(items)
        df_slip = df.groupby(['sku', 'title'], as_index=False) \
                   .agg({'quantity': 'sum', 'sku': 'first', 'orderKey': 'first'})
        df_slip.sort_values(by=['sku', 'title'], inplace=True)
        # Convert to list of dict
        slips = df_slip.to_dict(orient='records')
        # Convert to list of PickSlipItemVO
        slips = list(map(lambda x: PickSlipItemVO(**x), slips))
        return slips

    def sort_packing_order_refs(self, refs: List[str]) -> List[str]:
        """
        Sort the packing orders by the number of orderlines, orderKey and orderId
        :param refs: References of the orders to be sorted
        :return: List of references sorted by the criteria
        """
        pickItems = self.get_pick_items_by_references(refs)
        # Sort by [num_orderlines, orderKey and orderId]
        line_counts = count_elements(map(lambda x: x.orderId, pickItems))
        sku_counts = count_elements(map(lambda x: x.sku, pickItems))
        items = list(sorted(pickItems, key=lambda x: (line_counts[x.orderId],  x.orderKey, x.orderId)))
        sorted_refs = list(map(lambda x: x.orderId, items))
        return remove_duplicates(sorted_refs)

    def __add_excel_style(self, ws):
        # 遍历所有行，并根据orderId设置背景色
        gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        white_fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
        # 定义边框样式
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        alignment = Alignment(vertical="top", wrap_text=True)
        for row in range(2, ws.max_row + 1):  # 从第二行开始，因为第一行是表头
            fill = white_fill if row % 2 == 0 else gray_fill
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = alignment

        row_height = 50
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = row_height

    def pick_slip_to_excel(self, refs: List[str]) -> bytes:
        """
        TODO: Generate a pick slip excel file by references
        :param refs: List of references
        :return:
        """
        pickItems = self.make_pick_slip_with_references(refs)
        items = [o.dict() for o in pickItems]
        df_pick_slip = pd.DataFrame.from_dict(items)
        # Reduce columes
        df_pick_slip = df_pick_slip[['sku', 'quantity', 'storageLocation','title']]
        # Sort by sku
        df_pick_slip.sort_values(by=['sku'], inplace=True)
        # Convert dataframe to excel bytes
        with BytesIO() as excel_bytes:
            df_pick_slip.to_excel(excel_bytes, index=False, sheet_name="pick_slip")
            excel_bytes.seek(0)  # 重置文件指针
            wb = load_workbook(excel_bytes)
            ws = wb.active

            column_widths = [33, 10, 20, 80]  # 根据需要调整每列的宽度
            for col_num, width in enumerate(column_widths, start=1):
                col_letter = ws.cell(row=1, column=col_num).column_letter
                ws.column_dimensions[col_letter].width = width

            self.__add_excel_style(ws)

            # 将修改后的工作簿写回BytesIO对象
            new_excel_bytes = BytesIO()
            wb.save(new_excel_bytes)
            new_excel_bytes.seek(0)
            return new_excel_bytes.read()

    def pack_slips_to_excel(self, refs: List[str]) -> bytes:
        items = self.get_pick_items_by_references(refs)

        items = [o.dict() for o in items]
        df_order_lines = pd.DataFrame.from_dict(items)
        # Combine city and zip code to a single column
        df_order_lines['city'] = df_order_lines['zipCode'] + " " + df_order_lines['city']

        # Reduce columes
        df_order_lines = df_order_lines[["purchasedAt", 'taskId', 'orderId', 'sku', 'quantity', 'title', "city"]]
        # Format taskIds
        df_order_lines['taskId'] = df_order_lines['taskId'].apply(lambda x: f"{x: 03d}")
        # Format purchasedAt,
        df_order_lines['purchasedAt'] = df_order_lines['purchasedAt'] \
            .apply(lambda x: f"{datetime_to_date(x, target_pattern="%d.%m.%Y")}")

        with BytesIO() as excel_bytes:
            df_order_lines.to_excel(excel_bytes, index=False, sheet_name="pack_orders")
            excel_bytes.seek(0)  # 重置文件指针
            wb = load_workbook(excel_bytes)
            ws = wb.active

            column_widths = [13, 8, 25, 30, 10, 80, 30]  # 根据需要调整每列的宽度
            for col_num, width in enumerate(column_widths, start=1):
                col_letter = ws.cell(row=1, column=col_num).column_letter
                ws.column_dimensions[col_letter].width = width

            self.__add_excel_style(ws)

            # 将修改后的工作簿写回BytesIO对象
            new_excel_bytes = BytesIO()
            wb.save(new_excel_bytes)
            new_excel_bytes.seek(0)
            return new_excel_bytes.read()








